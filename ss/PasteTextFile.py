import pandas as pd
from openpyxl import Workbook

# File paths
txt_file_path = "C:\Temp\TTS\AM.txt"  # Path to your .txt file
excel_file_path = "C:\Temp\TTS\AMTTS.xlsx"  # Path to save the Excel file

with open(txt_file_path, "r") as file:
    lines = file.readlines()

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# Write lines to the sheet
row_idx = 1  # Start at row 1
split_mode = False  # Flag to identify when to split rows

for line in lines:
    stripped_line = line.strip()
    
    # Check for the header row to activate split mode
    if stripped_line.startswith("gta06_orig"):
        split_mode = True

    # Handle splitting for tabular data or headers
    if split_mode and stripped_line:
        # Split line into columns
        parts = stripped_line.split()
        for col_idx, value in enumerate(parts, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    else:
        # Write the line as-is in column A
        ws.cell(row=row_idx, column=1, value=stripped_line)
    
    row_idx += 1  # Increment row index

# Save the Excel file
wb.save(excel_file_path)

print(f"Data successfully written to {excel_file_path}")
