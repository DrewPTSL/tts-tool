import streamlit as st
import streamlit_ext as ste
import pandas as pd
import folium
from polyline import decode
from geopy.distance import geodesic
import re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import requests
from streamlit_folium import st_folium
import io
import plotly.express as px
import matplotlib.pyplot as plt
import geopandas as gpd
from shapely.geometry import Point
from folium.plugins import Search
from concurrent.futures import ThreadPoolExecutor, as_completed
import streamlit.components.v1 as components
import time

# --- Selenium webscraper imports ---
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pathlib import Path
import tempfile
import os

@st.cache_data(show_spinner="Loading zone data...")
def load_zones_data(data_choice):
    """Load zones data based on selected year"""
    if data_choice == "2006 Zones":
        zones_df = pd.read_csv('2006Zones.csv')
        zone_col = 'gta06'
        region_col = 'region'
    else:
        zones_df = pd.read_csv('2022Zones.csv')
        zone_col = 'TTS2022'
        region_col = 'Reg_name'
    return zones_df, zone_col, region_col

@st.cache_data(show_spinner="Loading polygons data...")
def load_geojson_data(data_choice):
    """Load GeoJSON data based on selected year"""
    if data_choice == "2006 Zones":
        file_path = "2006Polygons.geojson"
    else:
        file_path = "2022Polygons.geojson"
        
    gdf = gpd.read_file(file_path)
    if gdf.crs.to_epsg() != 4326:
        gdf = gdf.to_crs(epsg=4326)
    return gdf


# --- TTS Portal webscraper ------------------------------------------------
#
# Logs into the TTS cross-tabulation portal, runs one query per requested
# time period, and returns the combined raw text content as a single string
# (instead of leaving files on disk for manual download). This lets the
# scraped data flow directly into process_tts_file() alongside, or instead
# of, an uploaded .txt file.

def run_webscraper(site_zones, time_periods, data_choice, custom_time=None, headless=True):
    """
    Runs the TTS portal query for each requested time period and returns
    the concatenated raw text content as a single string, or None on failure.
    """
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    download_dir = str(Path.home() / "Downloads")
    chrome_options.add_experimental_option("prefs", {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    })

    if headless:
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)

    status_container = st.empty()
    progress_bar = st.progress(0)

    def update_status(message):
        status_container.text(message)

    combined_content_parts = []
    driver = None

    try:
        driver = webdriver.Chrome(
            service=Service("/usr/bin/chromedriver"),
            options=chrome_options
        )

        # Login process
        update_status("Logging into TTS system...")
        driver.get("https://drs.dmg.utoronto.ca/idrs/drsQuery/tts")
        time.sleep(1)

        username_field = driver.find_element(By.ID, "username")
        username_field.send_keys(st.secrets["USERNAME"])

        password_field = driver.find_element(By.ID, "password")
        password_field.send_keys(st.secrets["PASSWORD"])

        send_button = driver.find_element(By.ID, "send")
        send_button.click()
        time.sleep(1)

        # Set zone type based on data_choice
        if data_choice == "2006 Zones":
            origin_zone_type = "2006 GTA zone of origin"
            dest_zone_type = "2006 GTA zone of destination"
        else:
            origin_zone_type = "2022 TTS zone of origin"
            dest_zone_type = "2022 TTS zone of destination"

        zones_str = ", ".join(str(zone) for zone in site_zones)

        # Process time periods
        time_ranges = []
        for period in time_periods:
            if period == "AM Peak":
                time_ranges.append("700-930")
            elif period == "PM Peak":
                time_ranges.append("1600-1830")
            elif period == "All Day":
                time_ranges.append("400-2800")
            elif period == "Other" and custom_time:
                custom_ranges = [t.strip() for t in custom_time.split(',')]
                time_ranges.extend(custom_ranges)

        if not time_ranges:
            st.error("No valid time ranges to query.")
            return None

        total_iterations = len(time_ranges)
        current_iteration = 0

        # Navigate to cross tabulation page
        driver.get("https://drs.dmg.utoronto.ca/idrs/ttsForm/Cros/trip/2022")
        time.sleep(1)

        update_status("Setting up query parameters...")

        # Row variable
        row_variable = driver.find_element(By.XPATH, "//span[text()='Pick a Row Attribute']")
        row_variable.click()
        driver.switch_to.active_element.send_keys(origin_zone_type + Keys.RETURN)
        time.sleep(0.1)

        # Column variable
        column_variable = driver.find_element(By.XPATH, "//span[text()='Pick a Column Attribute']")
        column_variable.click()
        driver.switch_to.active_element.send_keys(dest_zone_type + Keys.RETURN)
        time.sleep(0.1)

        # Add filters
        add_button = driver.find_element(By.CLASS_NAME, "add")

        add_button.click()
        filter_1 = driver.find_element(By.XPATH, "//span[text()='Regional municipality of household']")
        filter_1.click()
        driver.switch_to.active_element.send_keys(origin_zone_type + Keys.RETURN)
        time.sleep(0.1)

        add_button.click()
        filter_2 = driver.find_element(By.XPATH, "//span[text()='Regional municipality of household']")
        filter_2.click()
        driver.switch_to.active_element.send_keys(dest_zone_type + Keys.RETURN)
        time.sleep(0.1)

        add_button.click()
        filter_3 = driver.find_element(By.XPATH, "//span[text()='Regional municipality of household']")
        filter_3.click()
        driver.switch_to.active_element.send_keys("Start time of trip" + Keys.RETURN)
        time.sleep(0.1)

        zone_textboxes = driver.find_elements(
            By.XPATH, '//input[@class="valuehtml ui-autocomplete-input" and @style="width:300px"]'
        )

        if len(zone_textboxes) >= 3:
            zone_textboxes[0].send_keys(str(zones_str))
            zone_textboxes[1].send_keys(str(zones_str))
            time.sleep(0.1)

        operator = driver.find_element(By.XPATH, "//span[text()='And']")
        operator.click()
        driver.switch_to.active_element.send_keys("Or" + Keys.RETURN)
        time.sleep(0.1)

        checkboxes = driver.find_elements(By.XPATH, '//input[@type="checkbox" and @class="toggle"]')
        if len(checkboxes) >= 3:
            checkboxes[-3].click()
            checkboxes[-2].click()
            time.sleep(0.1)

        radio_button = driver.find_element(By.ID, "emmeFormat")
        radio_button.click()

        for time_range in time_ranges:
            update_status(f"Processing zone{'' if len(site_zones) == 1 else 's'} {zones_str} for time period {time_range}")

            if len(zone_textboxes) >= 3:
                zone_textboxes[2].clear()
                time.sleep(0.1)
                zone_textboxes[2].send_keys(time_range)
                time.sleep(0.1)

            time.sleep(1)
            update_status(f"Executing query for time period {time_range} and downloading results...")
            execute_button = driver.find_element(By.CLASS_NAME, "submitCrosstab")
            execute_button.click()

            try:
                WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//button[@class='submitCrosstab' and text()='Execute Query']")
                    )
                )
                save_button = driver.find_element(By.CLASS_NAME, "saveAs")
                save_button.click()

                time.sleep(1)

                list_of_files = Path(download_dir).glob('*')
                latest_file = max(list_of_files, key=os.path.getctime)

                with open(latest_file, 'rb') as f:
                    file_content = f.read()

                # Decode and accumulate instead of leaving on disk for
                # separate manual downloads — this is what lets the scraped
                # data feed directly into process_tts_file().
                decoded = file_content.decode(errors='ignore')
                combined_content_parts.append(decoded)

                update_status(f"Downloaded data for time period {time_range}!")

            except Exception as e:
                st.error(f"Error downloading for zones {zones_str}, time {time_range}: {str(e)}")
                continue

            current_iteration += 1
            progress_bar.progress(current_iteration / total_iterations)

        update_status("All time periods processed successfully!")

        if not combined_content_parts:
            st.error("No data was successfully retrieved from any time period.")
            return None

        return "\n".join(combined_content_parts)

    except Exception as e:
        st.error(f"Error occurred while fetching TTS data: {str(e)}")
        return None
    finally:
        if driver is not None:
            driver.quit()


# Set page config
st.set_page_config(
    page_title="TTS Route Analysis Tool",
    page_icon="🚗",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.sidebar.title("🚗 TTS Route Analysis Tool")

# Initialize session states
if 'pois' not in st.session_state:
    st.session_state.pois = []
    
if 'poi_count' not in st.session_state:
    st.session_state.poi_count = 0
    
if 'processing_started' not in st.session_state:
    st.session_state.processing_started = False
    
if 'results_df' not in st.session_state:
    st.session_state.results_df = None

# Initialize rows if not in session_state
if "rows" not in st.session_state:
    st.session_state.rows = [{"id": 0, "name": "", "coords": "", "threshold": 50}]
if "row_id_counter" not in st.session_state:
    st.session_state.row_id_counter = 1

if "site_zones_val" not in st.session_state:
    st.session_state["site_zones_val"] = []

# Holds raw text fetched directly from the TTS portal, as an alternative
# to manually uploading a .txt file.
if "fetched_tts_content" not in st.session_state:
    st.session_state["fetched_tts_content"] = None

FOLIUM_TO_CSS = {
    'blue': '#4169E1',
    'red': '#DC143C',
    'green': '#228B22',
    'purple': '#800080',
    'orange': '#FF8C00',
    'darkred': '#8B0000',
    'lightred': '#FF6B6B',
    'beige': '#F5F5DC',
    'darkblue': '#00008B',
    'darkgreen': '#006400',
}

POI_COLOURS = ['blue', 'red', 'green', 'purple', 'orange', 'darkred',
               'lightred', 'beige', 'darkblue', 'darkgreen']

poi_colour_map = {
    poi['name']: POI_COLOURS[i % len(POI_COLOURS)]
    for i, poi in enumerate(st.session_state.pois)
}

# Title and description
st.title("TTS Route Analysis Tool")

# Add reset button in the top right
col1, col2 = st.columns([6, 1])

with col1:
    st.markdown("Upload your TTS file and analyze routes through points of interest.")

data_choice = st.radio(
    "Select Data Year:",
    options=["2006 Zones", "2022 Zones"],
    horizontal=True,
    index=None)

if data_choice:
    zones_df, zone_col, region_col = load_zones_data(data_choice)
    gdf = load_geojson_data(data_choice)
else:
    st.warning("Please select a data year")


# Configuration Section
st.markdown("### Site Configuration")
col1, col2 = st.columns(2)

with col1:
    if not data_choice:
        site_zones = st.multiselect("Site Zone", [], disabled=True)
    elif data_choice == "2006 Zones":
        site_zones = st.multiselect("Site Zone", zones_df['GTA06'], 
                                     default=st.session_state["site_zones_val"])
    else:
        site_zones = st.multiselect("Site Zone", zones_df['TTS2022'], 
                                     default=st.session_state["site_zones_val"])
    st.session_state["site_zones_val"] = site_zones

with col2:
    coords_input = st.text_input(
        "Site Coordinates (Latitude, Longitude)",
        value="",
        help="Enter coordinates in format: latitude, longitude"
    )

if coords_input.strip():  # Only validate if coordinates are provided
    try:
        site_lat, site_lon = map(float, coords_input.replace(" ", "").split(","))
        valid_coords = True
    except ValueError:
        st.error("Invalid coordinates format. Please use: latitude, longitude")
        valid_coords = False
else:
    valid_coords = False
    site_lat = None
    site_lon = None

## Site Zone Matching

if site_lon and data_choice:
    point = Point(site_lon, site_lat)
    matching_polygon = gdf[gdf.contains(point)]

    if not matching_polygon.empty:
        if data_choice == "2006 Zones":
            suggested_zone = matching_polygon.iloc[0]['gta06']
        else:
            suggested_zone = matching_polygon.iloc[0]['TTS2022']
        
        st.write(f"Recommended zone based on coordinates: {suggested_zone}")
        
        if st.button("➕ Add as Site Zone"):
            if suggested_zone not in st.session_state["site_zones_val"]:
                st.session_state["site_zones_val"].append(suggested_zone)
                st.rerun()
            else:
                st.info("Zone already added.")


# ---------------------------------------------------------------------------
# TTS Data Source Section: upload a file OR fetch directly from the portal
# ---------------------------------------------------------------------------
st.markdown("### TTS Data Source")

uploaded_file = st.file_uploader("Upload your TTS file", type=['txt'])

with st.expander("🌐 Or fetch directly from the TTS Portal"):
    if not data_choice:
        st.info("Select a data year above to enable fetching.")
    elif not site_zones:
        st.info("Select at least one Site Zone above to enable fetching.")
    else:
        time_period_options = ["AM Peak", "PM Peak", "All Day", "Other"]
        time_choice = st.pills(
            "Select Time Period(s):",
            time_period_options,
            selection_mode="single",
            key="fetch_time_periods"
        )

        custom_time = None
        if time_choice and "Other" in time_choice:
            custom_time = st.text_input(
                "Enter custom time range(s)",
                value="",
                help="e.g. 1200-1400 for 12 p.m. to 2 p.m. (separate multiple ranges with commas)",
                key="fetch_custom_time"
            )

        #headless = st.checkbox("Run browser headless", value=True, key="fetch_headless")

        if st.button("Fetch TTS Data", key="fetch_tts_button"):
            if not time_choice:
                st.error("Please select at least one time period.")
            else:
                with st.spinner("Fetching data from TTS portal..."):
                    fetched_content = run_webscraper(
                        site_zones=site_zones,
                        time_periods=time_choice,
                        data_choice=data_choice,
                        custom_time=custom_time,
                        headless=True
                    )
                if fetched_content:
                    st.session_state["fetched_tts_content"] = fetched_content
                    # A new uploaded file should always win if the user
                    # changes their mind later, so clear any stale results.
                    st.session_state.processing_started = False
                    st.session_state.results_df = None
                    st.success("TTS data fetched successfully — ready for analysis below.")
                    st.rerun()

    if st.session_state.get("fetched_tts_content"):
        st.success("✅ Fetched TTS data is loaded and will be used for analysis.")
        if st.button("Clear fetched data", key="clear_fetched_data"):
            st.session_state["fetched_tts_content"] = None
            st.rerun()


def get_tts_content():
    """
    Returns the raw TTS text content to process, preferring freshly
    uploaded files over previously fetched portal data, and the fetched
    data when no file has been uploaded.
    """
    if uploaded_file is not None:
        return uploaded_file.getvalue().decode()
    if st.session_state.get("fetched_tts_content"):
        return st.session_state["fetched_tts_content"]
    return None


has_tts_content = get_tts_content() is not None


# POI Management Section
st.markdown("### Points of Interest")

with st.expander("📋 Paste from Excel"):
    pasted = st.text_area(
        "Paste rows copied from Excel (expects columns: POI_ID, POI Name, Coordinates, Threshold (km))",
        height=150,
        placeholder="POI_ID\tPOI Name\tCoordinates\tThreshold (km)\nPOI_1\tNorth via West 5th Street\t43.2043, -79.8971\t0.05"
    )
    if st.button("Import"):
        if pasted.strip():
            new_rows = []
            lines = pasted.strip().split('\n')
            for line in lines:
                if line.lower().startswith('poi_id'):
                    continue
                parts = line.split('\t')
                if len(parts) >= 4:
                    name = parts[1].strip()
                    coords = parts[2].strip()
                    try:
                        threshold_m = int(float(parts[3].strip()) * 1000)
                    except ValueError:
                        threshold_m = 50
                    if name and coords:
                        new_rows.append({
                            "name": name,
                            "coords": coords,
                            "threshold": threshold_m
                        })
            
            if new_rows:
                existing = [row for row in st.session_state.rows if row["name"] and row["coords"]]
                for row in new_rows:
                    row["id"] = st.session_state.row_id_counter
                    st.session_state.row_id_counter += 1
                st.session_state.rows = existing + new_rows
                st.success(f"Imported {len(new_rows)} POIs. {len(existing)} existing POI(s) kept.")
                st.rerun()
            else:
                st.error("No valid rows found — make sure columns are tab-separated with POI_ID, POI Name, Coordinates, and Threshold (km).")

# Button to add a new row
if st.button("Add New Row"):
    st.session_state.rows.append({
        "id": st.session_state.row_id_counter,
        "name": "",
        "coords": "",
        "threshold": 50
    })
    st.session_state.row_id_counter += 1

num_rows = 1

# Display each row
for i, row in enumerate(st.session_state.rows):
    uid = row["id"]
    col1, col2, col3, col4 = st.columns([2, 2, 1, 0.5])
    with col1:
        row["name"] = st.text_input(
            "POI Name",
            key=f"name_{uid}",
            value=row["name"]
        )
    with col2:
        row["coords"] = st.text_input(
            "Coordinates (Latitude, Longitude)",
            key=f"coords_{uid}",
            value=row["coords"]
        )
    with col3:
        row["threshold"] = st.slider(
            "Threshold (m)",
            min_value=1,
            max_value=500,
            value=row["threshold"],
            key=f"threshold_{uid}"
        )
    with col4:
        if st.button("🗑️", key=f"delete_{uid}", help="Delete POI"):
            st.session_state.rows.pop(i)
            st.rerun()


# Process filled rows into POIs list before analysis
st.session_state.pois = []
for i, row in enumerate(st.session_state.rows):
    name = row["name"]
    coords = row["coords"]
    threshold = row["threshold"]
    
    if name and coords:
        try:
            lat, lon = map(float, coords.replace(" ", "").split(","))
            st.session_state.pois.append({
                'id': f'POI_{i + 1}',
                'name': name,
                'coordinates': (lat, lon),
                'threshold': threshold / 1000
            })
        except ValueError:
            st.error(f"Invalid coordinates format in row {i + 1}")

# After the site coordinates input section, add the map visualization
if valid_coords:
    try:          
        # Add a toggle button above the map
        show_map = st.toggle('Show Map', value=False)

        # Only display the map if toggle is on
        if show_map:
            m = folium.Map(location=[site_lat, site_lon], zoom_start=12, width='100%',tiles="CartoDB Voyager")

            # Add site zone marker
            sitezone_layer = folium.FeatureGroup(name="Site Zone Marker", show=True)
            folium.Marker(
                location=(site_lat, site_lon),
                popup=f"Site Zone {site_zones}",
                icon=folium.Icon(color='black', icon='home')
            ).add_to(sitezone_layer)

            # Add POIs with tooltips
            poi_layer = folium.FeatureGroup(name="POI Marker", show=True)
            poithreshold_layer = folium.FeatureGroup(name="POI Threshold Buffer", show=True)
            for poi in st.session_state.pois:
                folium.CircleMarker(
                    location=poi['coordinates'],
                    radius=5,
                    popup=f"POI ID: {poi['id']}<br>Name: {poi['name']}<br>Threshold: {poi['threshold']} km",
                    color='orange',
                    fill=True,
                    fillColor='orange',
                    fillOpacity=0.7
                ).add_to(poi_layer)

                # Add proximity circle for POIs with individual thresholds
                folium.Circle(
                    location=poi['coordinates'],
                    radius=poi['threshold'] * 1000,  # Convert km to meters
                    color='orange',
                    fill=True,
                    fillOpacity=0.2,
                    popup=f"{poi['name']}<br>Threshold: {poi['threshold']} km",
                ).add_to(poi_layer)

            def highlight_style_function(feature):
                return {
                    'fillColor': 'yellow',
                    'color': 'red',
                    'weight': 3,
                    'fillOpacity': 0.7
                }

            # Style function for regular polygons
            def style_function(feature):
                return {
                    'fillColor': 'white',
                    'color': 'black',
                    'weight': 2,
                    'fillOpacity': 0.5
                }

            # Highlighted polygon layer

            selectedzone_layer = folium.FeatureGroup(name="Selected Zones", show=True)
            for site_zone in site_zones:
                if site_zone == suggested_zone:
                    for _, row in matching_polygon.iterrows():
                        folium.GeoJson(
                            row.geometry,
                            style_function=style_function,
                            tooltip=f"{zone_col}: {row[zone_col]}, Region: {row[region_col]}"
                        ).add_to(selectedzone_layer)
                else:
                    for _, row in gdf.iterrows():
                        if row[zone_col] == site_zone:
                            folium.GeoJson(
                                row.geometry,
                                style_function=highlight_style_function,
                                tooltip=f"{zone_col}: {row[zone_col]}, Region: {row[region_col]}"
                            ).add_to(selectedzone_layer)
                


            # Add layer control
            selectedzone_layer.add_to(m)
            sitezone_layer.add_to(m)
            poi_layer.add_to(m)
            folium.LayerControl(collapsed=False).add_to(m)

            # Generate HTML for download AFTER adding all elements
            html = m.get_root().render()

            # Add download button AFTER map is fully configured
            ste.download_button(
                label="Download Map",
                data=html,
                file_name="site_poi_map.html",
                mime="text/html"
            )

            # Display map in Streamlit
            st.subheader("Site and POI Map")
            with st.container():
                st_folium(m, height=600, use_container_width=True, returned_objects=[])
        
    except Exception as e:
        st.error(f"Error creating map: {str(e)}")

## Main Processing Section
try:
    
    if not data_choice:
        st.warning("Please select a data year")
        st.stop()

    zones_df, zone_col, region_col = load_zones_data(data_choice)

    # zones_df, zone_col, region_col = load_zones_data(data_choice)

    
    # Validate site zones exist in zones.csv
    if all(zone in zones_df[zone_col].values for zone in site_zones) and valid_coords:
        if has_tts_content and len(st.session_state.pois) > 0:
            # Add start button
            start_button = st.button("Start Processing")
            
            if start_button:
                st.session_state.processing_started = True
            
            # Only show results if processing has started
            if st.session_state.processing_started:
                # Create a progress bar and status text
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                try:
                    def get_route(origin_lat, origin_lon, dest_lat, dest_lon, retries=3):
                        url = (f'http://router.project-osrm.org/route/v1/driving/'
                            f'{origin_lon},{origin_lat};{dest_lon},{dest_lat}?overview=full')
                        for attempt in range(retries):
                            try:
                                response = requests.get(url, timeout=10)
                                if response.status_code == 200:
                                    data = response.json()
                                    if data['code'] == 'Ok':
                                        return data['routes'][0]['geometry']
                            except requests.RequestException:
                                if attempt < retries - 1:
                                    time.sleep(1)
                        return None
                    
                    def fetch_routes_parallel(route_requests, max_workers=10, progress_callback=None, status_callback=None):
                        results = {}
                        total = len(route_requests)
                        completed = 0

                        with ThreadPoolExecutor(max_workers=max_workers) as executor:
                            futures = {
                                executor.submit(
                                    get_route,
                                    r['origin_lat'], r['origin_lon'],
                                    r['dest_lat'],   r['dest_lon']
                                ): r['key']
                                for r in route_requests
                            }
                            for future in as_completed(futures):
                                key = futures[future]
                                completed += 1
                                try:
                                    results[key] = future.result()
                                except Exception:
                                    results[key] = None

                                if progress_callback:
                                    # Fetching occupies 10% to 80% of the bar
                                    progress_callback(10 + int(70 * completed / total))
                                if status_callback:
                                    failed = sum(1 for v in results.values() if v is None)
                                    status_callback(
                                        f"Fetching routes... {completed} of {total} complete"
                                        + (f" ({failed} failed)" if failed > 0 else "")
                                    )

                        return results

                    def passes_through(route_geometry, poi_list, threshold=0.1):
                        route_coords = decode(route_geometry)
                        matched_pois = {}  # keyed by poi id — deduplicates automatically

                        for route_point in route_coords:
                            for poi in poi_list:
                                poi_id = poi['id']
                                if poi_id in matched_pois:
                                    continue  # already matched this POI, skip all future checks for it

                                distance = geodesic(route_point, poi['coordinates']).km
                                if distance <= poi.get('threshold', threshold):
                                    matched_pois[poi_id] = {
                                        'id': poi_id,
                                        'name': poi['name'],
                                        'coordinates': poi['coordinates'],
                                        'threshold': poi.get('threshold', threshold),
                                        'actual_distance': distance
                                    }

                            # Early exit if all POIs already matched
                            if len(matched_pois) == len(poi_list):
                                break

                        intersected = list(matched_pois.values())
                        return {
                            'passes': bool(intersected),
                            'num_pois_intersected': len(intersected),
                            'intersected_pois': intersected
                        }

                    def process_tts_file(content, zones_df, progress_callback=None, status_callback=None):
                        table_pattern = re.compile(r"^\s*(\d+)\s+(\d+)\s+(\d+)\s*$", re.MULTILINE)
                        matches = table_pattern.findall(content)
                        zone_col = 'GTA06' if data_choice == "2006 Zones" else 'TTS2022'
                        df_origins = pd.DataFrame(matches, columns=[f"{zone_col}_orig", f"{zone_col}_dest", "total"])
                        df_origins = df_origins.astype({f"{zone_col}_orig": int, f"{zone_col}_dest": int, "total": int})

                        zone_lookup = zones_df.set_index(zone_col)[['Latitude', 'Longitude']].to_dict('index')

                        # --- Phase 1: Plan routes ---
                        if status_callback:
                            status_callback(f"Planning routes for {len(site_zones)} site zone(s)...")
                        if progress_callback:
                            progress_callback(0)

                        route_requests = []
                        planned_rows = []

                        for current_site_zone in site_zones:
                            if current_site_zone not in zone_lookup:
                                continue
                            site_zone_coords = zone_lookup[current_site_zone]
                            szlat = site_zone_coords['Latitude']
                            szlon = site_zone_coords['Longitude']

                            for idx, row in df_origins.iterrows():
                                origin_id = row[f'{zone_col}_orig']
                                dest_id   = row[f'{zone_col}_dest']

                                if origin_id != current_site_zone and dest_id != current_site_zone:
                                    continue

                                origin_coords = zone_lookup.get(origin_id)
                                dest_coords   = zone_lookup.get(dest_id)

                                if not origin_coords or not dest_coords:
                                    planned_rows.append({
                                        'origin_id': origin_id,
                                        'dest_id': dest_id,
                                        'route_type': 'invalid_zone',
                                        'passes': False,
                                        'num_pois_intersected': 0,
                                        'intersected_pois': [],
                                        'total': row['total'],
                                        'site_zone': current_site_zone,
                                        'key': None,
                                        'geometry': None
                                    })
                                    continue

                                if origin_id == current_site_zone and dest_id == current_site_zone:
                                    key1 = f"{current_site_zone}|{idx}|origin_to_site"
                                    key2 = f"{current_site_zone}|{idx}|site_to_destination"
                                    route_requests.append({
                                        'key': key1,
                                        'origin_lat': szlat,
                                        'origin_lon': szlon,
                                        'dest_lat': site_lat,
                                        'dest_lon': site_lon
                                    })
                                    route_requests.append({
                                        'key': key2,
                                        'origin_lat': site_lat,
                                        'origin_lon': site_lon,
                                        'dest_lat': szlat,
                                        'dest_lon': szlon
                                    })
                                    planned_rows.append({
                                        'origin_id': origin_id,
                                        'dest_id': dest_id,
                                        'route_type': 'origin_to_site',
                                        'total': row['total'],
                                        'site_zone': current_site_zone,
                                        'key': key1,
                                        'geometry': None
                                    })
                                    planned_rows.append({
                                        'origin_id': origin_id,
                                        'dest_id': dest_id,
                                        'route_type': 'site_to_destination',
                                        'total': row['total'],
                                        'site_zone': current_site_zone,
                                        'key': key2,
                                        'geometry': None
                                    })

                                elif dest_id == current_site_zone:
                                    key = f"{current_site_zone}|{idx}|origin_to_site"
                                    route_requests.append({
                                        'key': key,
                                        'origin_lat': origin_coords['Latitude'],
                                        'origin_lon': origin_coords['Longitude'],
                                        'dest_lat': site_lat,
                                        'dest_lon': site_lon
                                    })
                                    planned_rows.append({
                                        'origin_id': origin_id,
                                        'dest_id': dest_id,
                                        'route_type': 'origin_to_site',
                                        'total': row['total'],
                                        'site_zone': current_site_zone,
                                        'key': key,
                                        'geometry': None
                                    })

                                else:
                                    key = f"{current_site_zone}|{idx}|site_to_destination"
                                    route_requests.append({
                                        'key': key,
                                        'origin_lat': site_lat,
                                        'origin_lon': site_lon,
                                        'dest_lat': dest_coords['Latitude'],
                                        'dest_lon': dest_coords['Longitude']
                                    })
                                    planned_rows.append({
                                        'origin_id': origin_id,
                                        'dest_id': dest_id,
                                        'route_type': 'site_to_destination',
                                        'total': row['total'],
                                        'site_zone': current_site_zone,
                                        'key': key,
                                        'geometry': None
                                    })

                        if status_callback:
                            status_callback(f"Found {len(route_requests)} routes to fetch — starting 10 workers...")
                        if progress_callback:
                            progress_callback(10)

                        # --- Phase 2: Fetch all routes in parallel ---
                        geometries = fetch_routes_parallel(
                            route_requests,
                            max_workers=10,
                            progress_callback=progress_callback,
                            status_callback=status_callback
                        )

                        # --- Phase 3: POI intersection checks ---
                        results = []
                        total = len(planned_rows)

                        for i, plan in enumerate(planned_rows):
                            if progress_callback:
                                progress_callback(80 + int(20 * i / total))
                            if status_callback:
                                status_callback(f"Checking POI intersections... {i+1} of {total}")

                            if plan.get('route_type') == 'invalid_zone' or plan['key'] is None:
                                results.append({
                                    'origin_id': plan['origin_id'],
                                    'dest_id': plan['dest_id'],
                                    'route_type': 'invalid_zone',
                                    'passes': False,
                                    'num_pois_intersected': 0,
                                    'intersected_pois': [],
                                    'total': plan['total'],
                                    'site_zone': plan['site_zone'],
                                    'geometry': None
                                })
                                continue

                            geometry = geometries.get(plan['key'])
                            if geometry:
                                poi_result = passes_through(geometry, st.session_state.pois)
                                results.append({
                                    'origin_id': plan['origin_id'],
                                    'dest_id': plan['dest_id'],
                                    'route_type': plan['route_type'],
                                    'passes': poi_result['passes'],
                                    'num_pois_intersected': poi_result['num_pois_intersected'],
                                    'intersected_pois': poi_result['intersected_pois'],
                                    'total': plan['total'],
                                    'site_zone': plan['site_zone'],
                                    'geometry': geometry
                                })
                            else:
                                results.append({
                                    'origin_id': plan['origin_id'],
                                    'dest_id': plan['dest_id'],
                                    'route_type': plan['route_type'],
                                    'passes': False,
                                    'num_pois_intersected': 0,
                                    'intersected_pois': [],
                                    'total': plan['total'],
                                    'site_zone': plan['site_zone'],
                                    'geometry': None
                                })

                        if status_callback:
                            status_callback("Processing complete!")
                        if progress_callback:
                            progress_callback(100)

                        st.session_state.zone_lookup = zone_lookup

                        return pd.DataFrame(results)

                    def update_progress(progress):
                        progress_bar.progress(int(progress))
                        
                    def update_status(status):
                        status_text.text(status)
                    
                    # Process the data — from upload or fetch (see get_tts_content)
                    content = get_tts_content()
                    st.session_state.results_df = process_tts_file(content, zones_df, update_progress, update_status)
                    
                    if st.session_state.results_df is not None and not st.session_state.results_df.empty:
                        status_text.text("Processing complete!")
                        
                        # Process the dataframe to show POI names
                        display_df = st.session_state.results_df.copy()
                        display_df['POI'] = display_df.apply(
                        lambda x: 'Invalid zone - route not processed' if x['route_type'] == 'invalid_zone' 
                        else (', '.join(sorted(set([poi['name'] for poi in x['intersected_pois']]))) if x['intersected_pois'] else ''),
                        axis=1
                        )
                        
                        # Select columns to display
                        display_df = display_df[['origin_id', 'dest_id', 'route_type', 'passes', 'POI', 'total']]
                        
                        # Display summary statistics
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("Total Routes", len(st.session_state.results_df))
                        with col2:
                            st.metric("Routes with POI Matches", st.session_state.results_df['passes'].sum())
                        
                        # Create POI summaries by route type
                        st.subheader("POI Traffic Distribution")
                        
                        # Filter for routes that pass through POIs
                        poi_df = display_df[display_df['POI'] != '']
                        
                        # Create two columns for the pie charts
                        col1, col2 = st.columns(2)
                        
                        # Split the data by route type
                        origin_to_site = poi_df[poi_df['route_type'] == 'origin_to_site']
                        site_to_destination = poi_df[poi_df['route_type'] == 'site_to_destination']
                        
                        # Create summary for origin_to_site
                        if not origin_to_site.empty:
                            with col1:
                                origin_summary = origin_to_site.groupby('POI')['total'].sum()
                                total_traffic = origin_summary.sum()
                                # Calculate percentages
                                origin_percentages = (origin_summary / total_traffic * 100).round(1)

                                poi_names_in_order = origin_summary.index.tolist()
                                
                                
                                # Create interactive pie chart
                                fig1 = px.pie(
                                    values=origin_percentages.values,
                                    names=origin_percentages.index,
                                    custom_data=[origin_summary.values],
                                    title="Origin to Site",
                                    color=origin_percentages.index,
                                    color_discrete_map={name: FOLIUM_TO_CSS.get(poi_colour_map.get(name, 'gray'), '#808080') 
                                                        for name in origin_percentages.index}
                                )
                                
                                fig1.update_traces(
                                    textposition='inside',
                                    hovertemplate="<b>%{label}</b><br>" +
                                                "Percentage: %{percent}<br>" +
                                                "Total Traffic: %{customdata[0]}<extra></extra>"
                                )
                                fig1.update_layout(
                                    showlegend=True,
                                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=1),
                                    height=400
                                )
                                st.plotly_chart(fig1, use_container_width=True)
                        
                        # Create summary for site_to_destination
                        if not site_to_destination.empty:
                            with col2:
                                dest_summary = site_to_destination.groupby('POI')['total'].sum()
                                total_traffic = dest_summary.sum()
                                # Calculate percentages
                                dest_percentages = (dest_summary / total_traffic * 100).round(1)

                                poi_names_in_order = dest_summary.index.tolist()
                                                                
                                # Create interactive pie chart
                                fig2 = px.pie(
                                    values=dest_percentages.values,
                                    names=dest_percentages.index,
                                    custom_data=[dest_summary.values],
                                    title="Site to Destination",
                                    color=dest_percentages.index,
                                    color_discrete_map={name: FOLIUM_TO_CSS.get(poi_colour_map.get(name, 'gray'), '#808080') 
                                                        for name in dest_percentages.index}
                                )
                                
                                fig2.update_traces(
                                    textposition='inside',
                                    hovertemplate="<b>%{label}</b><br>" +
                                                "Percentage: %{percent}<br>" +
                                                "Total Traffic: %{customdata[0]}<extra></extra>"
                                )
                                fig2.update_layout(
                                    showlegend=True,
                                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=1),
                                    height=400
                                )
                                st.plotly_chart(fig2, use_container_width=True)
                        
                        # Display results in a table
                        st.subheader("Route Analysis Results")
                        st.dataframe(display_df)
                        
                        # Generate Excel file for download
                        def generate_formatted_excel():
                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                # Format Route Results sheet
                                export_df = display_df[['origin_id', 'dest_id', 'route_type', 'POI', 'total']].copy()
                                export_df.to_excel(writer, sheet_name='Route Results', index=False)

                                # Create POI Summary DataFrame
                                poi_summary_df = pd.DataFrame([{
                                    'POI_ID': poi['id'],
                                    'POI Name': poi['name'],
                                    'Coordinates': f"{poi['coordinates'][0]}, {poi['coordinates'][1]}",
                                    'Threshold (km)': poi['threshold']
                                } for poi in st.session_state.pois])

                                # Create Site Summary DataFrame
                                site_summary_df = pd.DataFrame([{
                                    f'Site {zone_col} Zone': zone,
                                    'Zone Coordinates': f"{zones_df[zones_df[zone_col] == zone]['Latitude'].values[0]}, {zones_df[zones_df[zone_col] == zone]['Longitude'].values[0]}"
                                } for zone in site_zones])

                                site_location_summary_df = pd.DataFrame([{
                                    'Site Coordinates': f"{site_lat}, {site_lon}"
                                }])

                                # Write sheets
                                poi_summary_df.to_excel(writer, sheet_name='Location Details', startrow=0, startcol=0, index=False)
                                site_summary_df.to_excel(writer, sheet_name='Location Details', startrow=0, startcol=poi_summary_df.shape[1] + 2, index=False)
                                site_location_summary_df.to_excel(writer, sheet_name='Location Details', startrow=0, startcol=poi_summary_df.shape[1] + site_summary_df.shape[1] + 3, index=False)

                                # Apply Excel formatting
                                workbook = writer.book
                                for sheet_name in ['Route Results', 'Location Details']:
                                    sheet = writer.sheets[sheet_name]
                                    if sheet_name == 'Location Details':
                                        apply_header_formatting(sheet)
                                    else:
                                        apply_header_formatting(sheet)
                                    autofit_columns(sheet)

                                # Create and format POI Traffic Analysis sheet
                                calc_sheet = create_poi_analysis_sheet(workbook, st.session_state.pois)
                                format_poi_analysis_sheet(calc_sheet, len(st.session_state.pois))

                                # Create Raw Text sheet
                                create_raw_text_sheet(workbook, content)

                            return output.getvalue()

                        def apply_header_formatting(sheet, exclude_columns=None):
                            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                            header_fill = PatternFill(start_color='005295', end_color='005295', fill_type='solid')
                            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                        top=Side(style='thin'), bottom=Side(style='thin'))
                            if exclude_columns is None:
                                exclude_columns = []
                            for cell in sheet[1]:
                                if cell.column not in exclude_columns:
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.border = border
                                    cell.alignment = Alignment(horizontal='center')

                        def autofit_columns(sheet):
                            for column in sheet.columns:
                                max_length = 0
                                column = [cell for cell in column]
                                for cell in column:
                                    try:
                                        max_length = max(max_length, len(str(cell.value)))
                                    except:
                                        pass
                                adjusted_width = max_length + 2
                                sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

                        def create_poi_analysis_sheet(workbook, pois):
                            calc_sheet = workbook.create_sheet(title='POI Traffic Analysis')
                            total_row = len(pois) + 3
                            
                            # Non-rounded sums

                            calc_sheet['B2'] = 'From/To'
                            calc_sheet.merge_cells('C2:D2')
                            calc_sheet['C2'] = 'In'
                            calc_sheet.merge_cells('E2:F2')
                            calc_sheet['E2'] = 'Out'
                            
                            for idx, poi in enumerate(pois, start=3):
                                calc_sheet[f'B{idx}'] = poi['name']
                                calc_sheet[f'C{idx}'] = f'=SUMIFS(\'Route Results\'!E:E,\'Route Results\'!C:C,"origin_to_site",\'Route Results\'!D:D,"{poi["name"]}")'
                                calc_sheet[f'E{idx}'] = f'=SUMIFS(\'Route Results\'!E:E,\'Route Results\'!C:C,"site_to_destination",\'Route Results\'!D:D,"{poi["name"]}")'
                                calc_sheet[f'D{idx}'] = f'=IF(C{idx}>0,C{idx}/C{total_row},0)'
                                calc_sheet[f'F{idx}'] = f'=IF(E{idx}>0,E{idx}/E{total_row},0)'

                            calc_sheet[f'B{total_row}'] = "Total"
                            calc_sheet[f'C{total_row}'] = f'=SUM(C3:C{total_row-1})'
                            calc_sheet[f'D{total_row}'] = f'=SUM(D3:D{total_row-1})'
                            calc_sheet[f'E{total_row}'] = f'=SUM(E3:E{total_row-1})'
                            calc_sheet[f'F{total_row}'] = f'=SUM(F3:F{total_row-1})'

                            # Rounded sums

                            calc_sheet['H2'] = 'From/To'
                            calc_sheet['I2'] = 'In'
                            calc_sheet['J2'] = 'Out'

                            for idx, poi in enumerate(pois, start=3):
                                calc_sheet[f'H{idx}'] = poi['name']
                                calc_sheet[f'I{idx}'] = f'=MROUND(D{idx},0.05)'
                                calc_sheet[f'J{idx}'] = f'=MROUND(F{idx},0.05)'


                            calc_sheet[f'H{total_row}'] = "Total"
                            calc_sheet[f'I{total_row}'] = f'=SUM(I3:I{total_row-1})'
                            calc_sheet[f'J{total_row}'] = f'=SUM(J3:J{total_row-1})'
                            
                            return calc_sheet

                        def format_poi_analysis_sheet(sheet, num_pois):
                            total_row = num_pois + 3
                            
                            # Styles
                            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                            header_fill = PatternFill(start_color='005295', end_color='005295', fill_type='solid')
                            cell_font = Font(name='Arial', size=11)
                            total_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                            border_style = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                        top=Side(style='thin'), bottom=Side(style='thin'))

                            # Format headers
                            for col in ['B', 'C', 'D', 'E', 'F', 'H', 'I', 'J']:
                                cell = sheet[f'{col}2']
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = Alignment(horizontal='center')
                                cell.border = border_style

                            # Format POI rows
                            for row in range(3, total_row):
                                sheet[f'B{row}'].font = header_font
                                sheet[f'B{row}'].fill = header_fill
                                sheet[f'H{row}'].font = header_font
                                sheet[f'H{row}'].fill = header_fill
                                for col in ['C', 'D', 'E', 'F', 'I', 'J']:
                                    cell = sheet[f'{col}{row}']
                                    cell.font = cell_font
                                    cell.alignment = Alignment(horizontal='right')
                                    cell.border = border_style
                                    if col in ['D', 'F', 'I','J']:
                                        cell.number_format = '0.00%'

                            # Format the "Total" row
                            sheet[f'B{total_row}'].font = header_font
                            sheet[f'B{total_row}'].fill = header_fill
                            sheet[f'H{total_row}'].font = header_font
                            sheet[f'H{total_row}'].fill = header_fill

                            for col in ['C', 'D', 'E', 'F', 'I', 'J']:
                                total_cell = sheet[f'{col}{total_row}']
                                total_cell.font = Font(name='Arial', size=11, bold=True)
                                total_cell.fill = total_fill
                                total_cell.alignment = Alignment(horizontal='right')
                                total_cell.border = border_style
                                if col in ['D', 'F', 'I', 'J']:  # Ensure percentage format for Total row in columns D and F
                                    total_cell.number_format = '0.00%'

                            # Apply "all borders" to the entire table
                            for row in range(2, total_row + 1):
                                for col in ['B', 'C', 'D', 'E', 'F', 'H', 'I', 'J']:
                                    cell = sheet[f'{col}{row}']
                                    cell.border = border_style

                            # Autofit the width of column B
                            column_letters = ['B','H']
                            max_length = 0

                            # Iterate through all rows in column B to find the longest content
                            for row in sheet.iter_rows(min_col=2, max_col=2, min_row=2, max_row=total_row):
                                for cell in row:
                                    if cell.value:  # Ensure the cell has a value
                                        max_length = max(max_length, len(str(cell.value)))

                            # Adjust the column width (adding a little extra for padding)
                            for column_letter in column_letters:
                                sheet.column_dimensions[column_letter].width = max_length + 2

                        def create_raw_text_sheet(workbook, content):
                            raw_sheet = workbook.create_sheet(title='Raw Text')
                            row_idx = 1
                            split_mode = False
                            raw_sheet.sheet_view.show_grid_lines = False
                            
                            for line in content.split('\n'):
                                stripped_line = line.strip()
                                if stripped_line.startswith("gta06_orig") or stripped_line.startswith("tts22_orig"):
                                    split_mode = True
                                    
                                if split_mode and stripped_line:
                                    parts = stripped_line.split()
                                    for col_idx, value in enumerate(parts, start=1):
                                        raw_sheet.cell(row=row_idx, column=col_idx, value=value)
                                else:
                                    raw_sheet.cell(row=row_idx, column=1, value=stripped_line)
                                
                                row_idx += 1
                            
                        
                        excel_data = generate_formatted_excel()
                        ste.download_button(
                            label="Download Results as Excel",
                            data=excel_data,
                            file_name="tts_analysis_results.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                        if st.session_state.results_df is not None and not st.session_state.results_df.empty:
                            # Add a toggle button above the map
                            zone_lookup = st.session_state.get('zone_lookup', {})
                                # Only rebuild if results have changed
                            if 'route_map_html' not in st.session_state or \
                                st.session_state.get('route_map_results_id') != id(st.session_state.results_df):
                                    
                                    with st.spinner("Generating map..."):


                                        # Calculate max traffic for scaling route thickness
                                        max_traffic = st.session_state.results_df[
                                            st.session_state.results_df['passes']
                                        ]['total'].max() or 1

                                        def get_route_weight(total, max_traffic, min_weight=1, max_weight=8):
                                            """Scale route thickness between min and max weight based on traffic volume"""
                                            return min_weight + (max_weight - min_weight) * (total / max_traffic)

                                        route_map = folium.Map(location=[site_lat, site_lon], zoom_start=10,tiles="CartoDB Voyager")

                                        # Create feature groups
                                        site_layer = folium.FeatureGroup(name="Site Location", show=True)
                                        poi_layer = folium.FeatureGroup(name="Points of Interest", show=True)
                                        origin_nodes = folium.FeatureGroup(name="Origin Zone Locations", show=True)
                                        dest_nodes = folium.FeatureGroup(name="Destination Zone Locations", show=True)

                                        # Create one feature group per POI for origin and destination routes
                                        # so users can toggle each POI's routes independently
                                        origin_route_groups = {
                                            poi['name']: folium.FeatureGroup(
                                                name=f"Origin → Site via {poi['name']}", 
                                                show=True
                                            ) 
                                            for poi in st.session_state.pois
                                        }
                                        dest_route_groups = {
                                            poi['name']: folium.FeatureGroup(
                                                name=f"Site → Dest via {poi['name']}", 
                                                show=True
                                            ) 
                                            for poi in st.session_state.pois
                                        }

                                        # Use MarkerCluster for zone nodes to avoid clutter
                                        origin_cluster = folium.FeatureGroup(name="Origin Zone Locations", show=True)
                                        dest_cluster = folium.FeatureGroup(name="Destination Zone Locations", show=True)

                                        # Add site marker
                                        folium.Marker(
                                            location=[site_lat, site_lon],
                                            popup=folium.Popup("Site Location", max_width=200),
                                            icon=folium.Icon(color='black', icon='home')
                                        ).add_to(site_layer)

                                        # Track traffic totals per POI for the legend
                                        poi_traffic_in  = {poi['name']: 0 for poi in st.session_state.pois}
                                        poi_traffic_out = {poi['name']: 0 for poi in st.session_state.pois}

                                        # Add routes from stored geometry
                                        for _, row in st.session_state.results_df.iterrows():
                                            if not row['passes'] or row['geometry'] is None:
                                                continue

                                            coords = decode(row['geometry'])
                                            weight = get_route_weight(row['total'], max_traffic)

                                            # Get the first matched POI name to determine colour and group
                                            poi_name = row['intersected_pois'][0]['name'] if row['intersected_pois'] else None
                                            if not poi_name:
                                                continue

                                            colour = poi_colour_map.get(poi_name, 'gray')

                                            if row['route_type'] == 'origin_to_site':
                                                poi_traffic_in[poi_name] = poi_traffic_in.get(poi_name, 0) + row['total']
                                                folium.PolyLine(
                                                    coords,
                                                    weight=weight,
                                                    color=colour,
                                                    opacity=0.7,
                                                    popup=folium.Popup(
                                                        f"<b>Origin Zone:</b> {row['origin_id']}<br>"
                                                        f"<b>POI:</b> {poi_name}<br>"
                                                        f"<b>Traffic:</b> {row['total']}",
                                                        max_width=200
                                                    )
                                                ).add_to(origin_route_groups[poi_name])

                                            elif row['route_type'] == 'site_to_destination':
                                                poi_traffic_out[poi_name] = poi_traffic_out.get(poi_name, 0) + row['total']
                                                folium.PolyLine(
                                                    coords,
                                                    weight=weight,
                                                    color=colour,
                                                    opacity=0.7,
                                                    popup=folium.Popup(
                                                        f"<b>Destination Zone:</b> {row['dest_id']}<br>"
                                                        f"<b>POI:</b> {poi_name}<br>"
                                                        f"<b>Traffic:</b> {row['total']}",
                                                        max_width=200
                                                    )
                                                ).add_to(dest_route_groups[poi_name])

                                        # Add POI markers and threshold circles
                                        for poi in st.session_state.pois:
                                            colour = poi_colour_map[poi['name']]
                                            folium.CircleMarker(
                                                location=poi['coordinates'],
                                                radius=8,
                                                popup=folium.Popup(
                                                    f"<b>{poi['name']}</b><br>"
                                                    f"Threshold: {poi['threshold']} km<br>"
                                                    f"Traffic In: {poi_traffic_in.get(poi['name'], 0)}<br>"
                                                    f"Traffic Out: {poi_traffic_out.get(poi['name'], 0)}",
                                                    max_width=200
                                                ),
                                                color=colour,
                                                fill=True,
                                                fillColor=colour,
                                                fillOpacity=0.9
                                            ).add_to(poi_layer)

                                            folium.Circle(
                                                location=poi['coordinates'],
                                                radius=poi['threshold'] * 1000,
                                                color=colour,
                                                fill=True,
                                                fillOpacity=0.15,
                                                popup=f"{poi['name']}<br>Threshold: {poi['threshold']} km",
                                            ).add_to(poi_layer)

                                        # Add zone node markers to clusters
                                        for _, row in st.session_state.results_df.iterrows():
                                            if not row['passes']:
                                                continue

                                            poi_name = row['intersected_pois'][0]['name'] if row['intersected_pois'] else 'Unknown'
                                            colour = poi_colour_map.get(poi_name, 'gray')

                                            if row['route_type'] == 'origin_to_site':
                                                zone_id = row['origin_id']
                                                zone_row = zone_lookup.get(zone_id)
                                                if zone_row:
                                                    folium.Marker(
                                                        location=[zone_row['Latitude'], zone_row['Longitude']],
                                                        popup=folium.Popup(
                                                            f"<b>Origin Zone:</b> {zone_id}<br>"
                                                            f"<b>POI:</b> {poi_name}<br>"
                                                            f"<b>Total Trips:</b> {row['total']}",
                                                            max_width=200
                                                        ),
                                                        icon=folium.Icon(color=colour, icon='car', prefix='fa')
                                                    ).add_to(origin_cluster)

                                            else:
                                                zone_id = row['dest_id']
                                                zone_row = zone_lookup.get(zone_id)
                                                if zone_row:
                                                    folium.Marker(
                                                        location=[zone_row['Latitude'], zone_row['Longitude']],
                                                        popup=folium.Popup(
                                                            f"<b>Destination Zone:</b> {zone_id}<br>"
                                                            f"<b>POI:</b> {poi_name}<br>"
                                                            f"<b>Total Trips:</b> {row['total']}",
                                                            max_width=200
                                                        ),
                                                        icon=folium.Icon(color=colour, icon='car-side', prefix='fa')
                                                    ).add_to(dest_cluster)

                                        # Build legend HTML
                                        legend_html = """
                                        <div style="position: fixed; bottom: 40px; left: 40px; z-index: 1000;
                                                    background-color: white; padding: 12px 16px; border-radius: 8px;
                                                    border: 1px solid #ccc; font-family: Arial; font-size: 12px;
                                                    box-shadow: 2px 2px 6px rgba(0,0,0,0.2); min-width: 200px;">
                                            <b style="font-size:13px;">POI Traffic Summary</b><br><br>
                                        """
                                        for poi in st.session_state.pois:
                                            colour = poi_colour_map[poi['name']]
                                            traffic_in  = poi_traffic_in.get(poi['name'], 0)
                                            traffic_out = poi_traffic_out.get(poi['name'], 0)
                                            total_in    = sum(poi_traffic_in.values()) or 1
                                            total_out   = sum(poi_traffic_out.values()) or 1
                                            pct_in      = round(traffic_in / total_in * 100, 1)
                                            pct_out     = round(traffic_out / total_out * 100, 1)
                                            legend_html += f"""
                                            <div style="margin-bottom:6px;">
                                                <span style="display:inline-block; width:14px; height:14px; 
                                                            background:{colour}; border-radius:50%; 
                                                            margin-right:6px; vertical-align:middle;"></span>
                                                <b>{poi['name']}</b><br>
                                                <span style="margin-left:20px;">In: {traffic_in} ({pct_in}%)</span><br>
                                                <span style="margin-left:20px;">Out: {traffic_out} ({pct_out}%)</span>
                                            </div>
                                            """
                                        legend_html += "</div>"

                                        route_map.get_root().html.add_child(folium.Element(legend_html))

                                        # Add route weight legend
                                        weight_legend_html = """
                                        <div style="position: fixed; bottom: 40px; right: 40px; z-index: 1000;
                                                    background-color: white; padding: 12px 16px; border-radius: 8px;
                                                    border: 1px solid #ccc; font-family: Arial; font-size: 12px;
                                                    box-shadow: 2px 2px 6px rgba(0,0,0,0.2);">
                                            <b style="font-size:13px;">Route Thickness</b><br><br>
                                            <svg width="120" height="60">
                                                <line x1="0" y1="12" x2="120" y2="12" stroke="#555" stroke-width="1"/>
                                                <text x="0" y="26" font-size="10">Low traffic</text>
                                                <line x1="0" y1="44" x2="120" y2="44" stroke="#555" stroke-width="8"/>
                                                <text x="0" y="58" font-size="10">High traffic</text>
                                            </svg>
                                        </div>
                                        """
                                        route_map.get_root().html.add_child(folium.Element(weight_legend_html))

                                        # Add all layers to map
                                        site_layer.add_to(route_map)
                                        poi_layer.add_to(route_map)
                                        for group in origin_route_groups.values():
                                            group.add_to(route_map)
                                        for group in dest_route_groups.values():
                                            group.add_to(route_map)
                                        origin_cluster.add_to(route_map)
                                        dest_cluster.add_to(route_map)
                                        folium.LayerControl(collapsed=False).add_to(route_map)

                                        # Cache the rendered HTML
                                        st.session_state.route_map_html = route_map.get_root().render()
                                        st.session_state.route_map_results_id = id(st.session_state.results_df)

                                        # Cache the rendered HTML and a download copy
                                        st.session_state.route_map_html = route_map.get_root().render()
                                        st.session_state.route_map_results_id = id(st.session_state.results_df)

                            # Download button uses cached HTML
                            ste.download_button(
                                label="Download Route Map",
                                data=st.session_state.route_map_html,
                                file_name="Route_map.html",
                                mime="text/html"
                                )

                            st.subheader("Route Map")
                            components.html(st.session_state.route_map_html, height=600)

                except Exception as e:
                    status_text.text(f"Error during processing: {str(e)}")
                    progress_bar.progress(0)
                    st.error(f"An error occurred: {str(e)}")

        elif len(st.session_state.pois) == 0:
            st.warning("Please add at least one Point of Interest before processing")
        else:
            st.warning("Please upload a TTS file or fetch data from the portal to process")
    else:
        if not valid_coords:
            st.warning("Please enter valid Site Coordinates")
        else:
            st.error("Site zone does not exist in zones.csv")
except Exception as e:
    st.error(f"An error occurred: {str(e)}")