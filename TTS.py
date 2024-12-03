import requests
import pandas as pd
import folium
from polyline import decode
from geopy.distance import geodesic
import re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook

# Step 1: Read the TTS Text File and extract GTA06 origins
with open(r"TTSAM.txt", "r") as file:
    content = file.read()

# Re-open file to get lines
with open(r"TTSAM.txt", "r") as file:
    lines = file.readlines()

# Extract table from text file
table_pattern = re.compile(r"^\s*(\d+)\s+(\d+)\s+(\d+)\s*$", re.MULTILINE)
matches = table_pattern.findall(content)
df_origins = pd.DataFrame(matches, columns=["gta06_orig", "gta06_dest", "total"])
df_origins = df_origins.astype({"gta06_orig": int, "gta06_dest": int, "total": int})

# Read zones CSV
zones_df = pd.read_csv(r'Zones.csv')

# Merge to get latitude and longitude for origins and destinations
origins_df = zones_df.merge(df_origins, left_on='GTA06', right_on='gta06_orig')
destinations_df = zones_df.merge(df_origins, left_on='GTA06', right_on='gta06_dest')


# Your site coordinates and zone
site_coords = "43.212342441568325, -79.77345451007199"  # Format: "latitude, longitude"
site_lat, site_lon = map(float, site_coords.split(','))

site_zone = 5241

# Points of interest (POIs) to check
pois = [
    {
        'id': 'POI_1',
        'name': 'South via Greenhill Avenue',
        'coordinates': (43.21107989319771, -79.77681496454343),
        'threshold': 0.05  # 50 meters
    },
    {
        'id': 'POI_2', 
        'name': 'North via Greenhill Avenue',
        'coordinates': (43.21516364199378, -79.7747902086894),
        'threshold': 0.1   # 100 meters
    }
]

def get_route(origin_lat, origin_lon, dest_lat, dest_lon):
    url = f'http://router.project-osrm.org/route/v1/driving/{origin_lon},{origin_lat};{dest_lon},{dest_lat}?overview=full'
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        if data['code'] == 'Ok':
            return data['routes'][0]['geometry']
    return None

def passes_through(route_geometry, poi_list, threshold=0.1):
    route_coords = decode(route_geometry)
    matching_pois = []
    
    for route_point in route_coords:
        for poi in poi_list:
            poi_threshold = poi.get('threshold', threshold)  # Use POI-specific threshold or default
            distance = geodesic(route_point, poi['coordinates']).km
            if distance <= poi_threshold:
                matching_pois.append({
                    'id': poi['id'],
                    'name': poi['name'],
                    'coordinates': poi['coordinates'],
                    'threshold': poi_threshold,
                    'actual_distance': distance
                })
    
    # Remove duplicate POIs
    matching_pois = [dict(t) for t in {tuple(d.items()) for d in matching_pois}]
    
    return {
        'passes': bool(matching_pois),
        'num_pois_intersected': len(matching_pois),
        'intersected_pois': matching_pois
    }

# Get routes and check for POIs
routes = []
results = []

# Get site zone coordinates
# Add error checking for site zone
if zones_df[zones_df['GTA06'] == site_zone].empty:
    raise ValueError(f"Site zone {site_zone} not found in Zones.csv")

try:
    site_zone_row = zones_df[zones_df['GTA06'] == site_zone]
    site_zone_lat = site_zone_row['Latitude'].values[0]
    site_zone_lon = site_zone_row['Longitude'].values[0]
except IndexError:
    raise ValueError(f"Could not retrieve coordinates for site zone {site_zone}")

# Debug prints
print("Site Zone:", site_zone)
print("Total Origins/Destinations:", len(df_origins))
print("Origins to Site Zone:", len(df_origins[df_origins['gta06_dest'] == site_zone]))
print("Origins from Site Zone:", len(df_origins[df_origins['gta06_orig'] == site_zone]))

# Check if there are any routes involving the site zone
site_routes = df_origins[(df_origins['gta06_orig'] == site_zone) | (df_origins['gta06_dest'] == site_zone)]
print("Routes involving site zone:", len(site_routes))

# Print out unique zones in the dataframe
print("Unique Zones in Origins/Destinations:", df_origins['gta06_orig'].unique())
print("Unique Zones in Origins/Destinations:", df_origins['gta06_dest'].unique())


for index, row in df_origins.iterrows():
    origin_id = row['gta06_orig']
    dest_id = row['gta06_dest']
    current_route = None
    # Debugging: Print current route details
    print(f"Processing route: Origin {origin_id}, Destination {dest_id}")
    
    # Check if origin zone exists in zones_df
    origin_row = zones_df[zones_df['GTA06'] == origin_id]
    if origin_row.empty:
        print(f"WARNING: Origin zone {origin_id} not found in Zones.csv")
        results.append({
            'origin_id': origin_id,
            'dest_id': dest_id,
            'route_type': 'origin_to_site' if dest_id == site_zone else 'unknown',
            'passes': False,
            'num_pois_intersected': 0,
            'intersected_pois_ids': [f'Error: Origin zone {origin_id} not found in Zones.csv'],
            'intersected_pois_names': [],
            'intersected_pois_details': []
        })
        continue  # Skip route processing but still include in results
    
    # Check if destination zone exists in zones_df
    dest_row = zones_df[zones_df['GTA06'] == dest_id]
    if dest_row.empty:
        print(f"WARNING: Destination zone {dest_id} not found in Zones.csv")
        results.append({
            'origin_id': origin_id,
            'dest_id': dest_id,
            'route_type': 'site_to_destination' if origin_id == site_zone else 'unknown',
            'passes': False,
            'num_pois_intersected': 0,
            'intersected_pois_ids': [f'Error: Destination zone {dest_id} not found in Zones.csv'],
            'intersected_pois_names': [],
            'intersected_pois_details': []
        })
        continue  # Skip route processing but still include in results
    
    # Safely extract origin coordinates
    try:
        origin_lat = origin_row['Latitude'].values[0]
        origin_lon = origin_row['Longitude'].values[0]
    except IndexError:
        print(f"ERROR: Could not extract coordinates for origin zone {origin_id}")
        results.append({
            'origin_id': origin_id,
            'dest_id': dest_id,
            'route_type': 'origin_to_site' if dest_id == site_zone else 'unknown',
            'passes': False,
            'num_pois_intersected': 0,
            'intersected_pois_ids': [f'Error: Could not extract coordinates for origin zone {origin_id}'],
            'intersected_pois_names': [],
            'intersected_pois_details': []
        })
        continue
        
    # Safely extract destination coordinates
    try:
        dest_lat = dest_row['Latitude'].values[0]
        dest_lon = dest_row['Longitude'].values[0]
    except IndexError:
        print(f"ERROR: Could not extract coordinates for destination zone {dest_id}")
        results.append({
            'origin_id': origin_id,
            'dest_id': dest_id,
            'route_type': 'site_to_destination' if origin_id == site_zone else 'unknown',
            'passes': False,
            'num_pois_intersected': 0,
            'intersected_pois_ids': [f'Error: Could not extract coordinates for destination zone {dest_id}'],
            'intersected_pois_names': [],
            'intersected_pois_details': []
        })
        continue
    
    # Case 1: Both origin and destination are the site zone (highest priority)
    if origin_id == site_zone and dest_id == site_zone:
        # Route from site zone to site (origin_to_site)
        route1 = get_route(site_zone_lat, site_zone_lon, site_lat, site_lon)
        
        # Route from site to site zone (site_to_destination)
        route2 = get_route(site_lat, site_lon, site_zone_lat, site_zone_lon)
        
        routes.extend([route1, route2])
        
        # Process origin_to_site route
        if route1:
            poi_check_result = passes_through(route1, pois, threshold=0.05)
            results.append({
                'origin_id': origin_id, 
                'dest_id': dest_id,
                'route_type': 'origin_to_site',
                'passes': poi_check_result['passes'], 
                'num_pois_intersected': poi_check_result['num_pois_intersected'],
                'intersected_pois_ids': [poi['id'] for poi in poi_check_result['intersected_pois']],
                'intersected_pois_names': [poi['name'] for poi in poi_check_result['intersected_pois']],
                'intersected_pois_details': poi_check_result['intersected_pois']
            })
        else:
            results.append({
                'origin_id': origin_id, 
                'dest_id': dest_id,
                'route_type': 'origin_to_site',
                'passes': False, 
                'num_pois_intersected': 0,
                'intersected_pois_ids': [],
                'intersected_pois_names': [],
                'intersected_pois_details': []
            })

        # Process site_to_destination route
        if route2:
            poi_check_result = passes_through(route2, pois, threshold=0.05)
            results.append({
                'origin_id': origin_id, 
                'dest_id': dest_id,
                'route_type': 'site_to_destination',
                'passes': poi_check_result['passes'], 
                'num_pois_intersected': poi_check_result['num_pois_intersected'],
                'intersected_pois_ids': [poi['id'] for poi in poi_check_result['intersected_pois']],
                'intersected_pois_names': [poi['name'] for poi in poi_check_result['intersected_pois']],
                'intersected_pois_details': poi_check_result['intersected_pois']
            })
        else:
            results.append({
                'origin_id': origin_id, 
                'dest_id': dest_id,
                'route_type': 'site_to_destination',
                'passes': False, 
                'num_pois_intersected': 0,
                'intersected_pois_ids': [],
                'intersected_pois_names': [],
                'intersected_pois_details': []
            })
        continue  # Skip to next iteration after processing routes
    
    # Case 2: Origin is not the site zone (route from origin to site)
    if origin_id != site_zone and dest_id == site_zone:
        origin_row = zones_df[zones_df['GTA06'] == origin_id]
        origin_lat = origin_row['Latitude'].values[0]
        origin_lon = origin_row['Longitude'].values[0]
        
        current_route = get_route(origin_lat, origin_lon, site_lat, site_lon)
        
        if current_route:
            routes.append(current_route)
            poi_check_result = passes_through(current_route, pois, threshold=0.05)
            results.append({
                'origin_id': origin_id, 
                'dest_id': dest_id,
                'route_type': 'origin_to_site',
                'passes': poi_check_result['passes'], 
                'num_pois_intersected': poi_check_result['num_pois_intersected'],
                'intersected_pois_ids': [poi['id'] for poi in poi_check_result['intersected_pois']],
                'intersected_pois_names': [poi['name'] for poi in poi_check_result['intersected_pois']],
                'intersected_pois_details': poi_check_result['intersected_pois']
            })
        else:
            routes.append(None)
            results.append({
                'origin_id': origin_id, 
                'dest_id': dest_id,
                'route_type': 'origin_to_site',
                'passes': False, 
                'num_pois_intersected': 0,
                'intersected_pois_ids': [],
                'intersected_pois_names': [],
                'intersected_pois_details': []
            })
    
    # Case 3: Origin is the site zone (route from site to destination)
    if origin_id == site_zone and dest_id != site_zone:
        dest_row = zones_df[zones_df['GTA06'] == dest_id]
        dest_lat = dest_row['Latitude'].values[0]
        dest_lon = dest_row['Longitude'].values[0]
        
        current_route = get_route(site_lat, site_lon, dest_lat, dest_lon)
        
        if current_route:
            routes.append(current_route)
            poi_check_result = passes_through(current_route, pois, threshold=0.05)
            results.append({
                'origin_id': origin_id, 
                'dest_id': dest_id,
                'route_type': 'site_to_destination',
                'passes': poi_check_result['passes'], 
                'num_pois_intersected': poi_check_result['num_pois_intersected'],
                'intersected_pois_ids': [poi['id'] for poi in poi_check_result['intersected_pois']],
                'intersected_pois_names': [poi['name'] for poi in poi_check_result['intersected_pois']],
                'intersected_pois_details': poi_check_result['intersected_pois']
            })
        else:
            routes.append(None)
            results.append({
                'origin_id': origin_id, 
                'dest_id': dest_id,
                'route_type': 'site_to_destination',
                'passes': False, 
                'num_pois_intersected': 0,
                'intersected_pois_ids': [],
                'intersected_pois_names': [],
                'intersected_pois_details': []
            })

results_df = pd.DataFrame(results)

# Merge the total column from the original dataframe
results_df = results_df.merge(df_origins[['gta06_orig', 'gta06_dest', 'total']], 
                               left_on=['origin_id', 'dest_id'], 
                               right_on=['gta06_orig', 'gta06_dest'], 
                               how='left')

# Create POI Summary DataFrame
poi_summary_df = pd.DataFrame([
    {
        'Location Type': poi['name'], 
        'POI_ID': poi['id'], 
        'Latitude': poi['coordinates'][0],
        'Longitude': poi['coordinates'][1],
        'Threshold (km)': poi['threshold']
    } for poi in pois
])

# Convert lists to strings for Excel output, ensuring uniqueness
results_df['intersected_pois_ids'] = results_df['intersected_pois_ids'].apply(lambda x: ', '.join(list(dict.fromkeys(x))) if isinstance(x, list) else x)
results_df['intersected_pois_names'] = results_df['intersected_pois_names'].apply(lambda x: ', '.join(list(dict.fromkeys(x))) if isinstance(x, list) else x)

# Export to Excel
export_df = results_df[['origin_id', 'dest_id', 'route_type', 'intersected_pois_ids', 'intersected_pois_names', 'total']].copy()
export_df.columns = ['origin_id', 'dest_id', 'route_type', 'intersected_ids', 'intersected_names', 'total']

# Save to Excel
export_df.to_excel("Output.xlsx", index=False)
print("Results saved to Output.xlsx")

# Create POI Summary DataFrame
poi_summary_df = pd.DataFrame([
    {
        'Location Type': poi['name'], 
        'POI_ID': poi['id'], 
        'Latitude': poi['coordinates'][0],
        'Longitude': poi['coordinates'][1],
        'Threshold (km)': poi['threshold']
    } for poi in pois
])

# Create Site Summary DataFrame
site_summary_df = pd.DataFrame([{
    'Site Zone': site_zone,
    'Latitude': site_lat,
    'Longitude': site_lon
}, {
    'Site Zone': 'Additional Info',
    'Latitude': 'Placeholder for any extra site information',
    'Longitude': ''
}])

# Create an Excel writer
with pd.ExcelWriter('Output.xlsx', engine='openpyxl') as writer:
    # Write the main results sheet
    workbook = writer.book
    wb = Workbook()
    raw_text_sheet = workbook.create_sheet(title='Raw Text')

    row_idx = 1  # Start at row 1
    split_mode = False  # Flag to identify when to split rows

    for line in lines:
        stripped_line = line.strip()
        
        # Check for the header row to activate split mode
        if stripped_line.startswith("gta06_orig"):
            split_mode = True

        # Handle splitting for tabular data or headers
        if split_mode and stripped_line:
            # Split line into columns
            parts = stripped_line.split()
            for col_idx, value in enumerate(parts, start=1):
                raw_text_sheet.cell(row=row_idx, column=col_idx, value=value)
        else:
            # Write the line as-is in column A
            raw_text_sheet.cell(row=row_idx, column=1, value=stripped_line)
        
        row_idx += 1  # Increment row index

    export_df.to_excel(writer, sheet_name='Route Results', index=False)
   
    # Write the Location Details sheet
    poi_summary_df.to_excel(writer, sheet_name='Location Details', startrow=0, startcol=0, index=False)
    site_summary_df.to_excel(writer, sheet_name='Location Details', startrow=0, startcol=poi_summary_df.shape[1] + 2, index=False)
    
    # Get the workbook and create a new sheet for calculations
    workbook = writer.book
    calc_sheet = workbook.create_sheet(title='POI Traffic Analysis') 

    # Set up headers
    calc_sheet['B2'] = 'From/To'
    
    # Populate POI names
    for idx, poi in enumerate(pois, start=3):
        calc_sheet[f'B{idx}'] = poi['name']
    
    # Merge cells for In and Out columns
    calc_sheet.merge_cells('C2:D2')
    calc_sheet['C2'] = 'In'
    calc_sheet.merge_cells('E2:F2')
    calc_sheet['E2'] = 'Out'
    
    # Create formulas for each POI using SUMIFS
    for idx, poi in enumerate(pois, start=3):
        # Formula for In traffic (origin_to_site)
        calc_sheet[f'C{idx}'] = f'=SUMIFS(\'Route Results\'!F:F,\'Route Results\'!C:C,"origin_to_site",\'Route Results\'!D:D,"{poi["id"]}")'
        
        # Formula for Out traffic (site_to_destination)
        calc_sheet[f'E{idx}'] = f'=SUMIFS(\'Route Results\'!F:F,\'Route Results\'!C:C,"site_to_destination",\'Route Results\'!D:D,"{poi["id"]}")'

    # Adding the "Total" row below the last POI
    total_row = len(pois) + 3  # Calculate the row number for the "Total" row
    calc_sheet[f'B{total_row}'] = "Total"  # Label the "Total" row

    # Add the sum formulas for the "In" and "Out" columns
    calc_sheet[f'C{total_row}'] = f'=SUM(C3:C{total_row - 1})'
    calc_sheet[f'D{total_row}'] = f'=SUM(D3:D{total_row - 1})'
    calc_sheet[f'E{total_row}'] = f'=SUM(E3:E{total_row - 1})'
    calc_sheet[f'F{total_row}'] = f'=SUM(F3:F{total_row - 1})'


    # Add formulas for percentage of total in columns D and F
    for idx, poi in enumerate(pois, start=3):
        # Percentage for "In" traffic
        calc_sheet[f'D{idx}'] = f'=IF(C{idx}>0, C{idx}/C{total_row}, 0)'
        # Percentage for "Out" traffic
        calc_sheet[f'F{idx}'] = f'=IF(E{idx}>0, E{idx}/E{total_row}, 0)'

    # Format the percentage columns as percentages
    for row in range(3, total_row+1):
        calc_sheet[f'D{row}'].number_format = '0.00%'
        calc_sheet[f'F{row}'].number_format = '0.00%'

    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='005295', end_color='005295', fill_type='solid')
    cell_font = Font(name='Arial', size=11)
    total_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    alignment_right = Alignment(horizontal='right')
    border_style = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    # Format the header row
    for col in ['B', 'C', 'D', 'E', 'F']:
        header_cell = calc_sheet[f'{col}2']
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = Alignment(horizontal='center')
        header_cell.border = border_style

    # Format POI rows
    for row in range(3, total_row):
        calc_sheet[f'B{row}'].font = header_font  # Bold POI names
        calc_sheet[f'B{row}'].fill = header_fill
        for col in ['C', 'D', 'E', 'F']:
            cell = calc_sheet[f'{col}{row}']
            cell.font = cell_font
            cell.alignment = alignment_right
            cell.border = border_style
            if col in ['D', 'F']:  # Percentages
                cell.number_format = '0.00%'

    # Format the "Total" row
    calc_sheet[f'B{total_row}'].font = header_font
    calc_sheet[f'B{total_row}'].fill = header_fill

    for col in ['C', 'D', 'E', 'F']:
        total_cell = calc_sheet[f'{col}{total_row}']
        total_cell.font = Font(name='Arial', size=11, bold=True)
        total_cell.fill = total_fill
        total_cell.alignment = alignment_right
        total_cell.border = border_style

    # Apply "all borders" to the entire table
    for row in range(2, total_row + 1):
        for col in ['B', 'C', 'D', 'E', 'F']:
            cell = calc_sheet[f'{col}{row}']
            cell.border = border_style

    # Autofit the width of column B
    column_letter = 'B'
    max_length = 0

    # Iterate through all rows in column B to find the longest content
    for row in calc_sheet.iter_rows(min_col=2, max_col=2, min_row=2, max_row=total_row):
        for cell in row:
            if cell.value:  # Ensure the cell has a value
                max_length = max(max_length, len(str(cell.value)))

    # Adjust the column width (adding a little extra for padding)
    calc_sheet.column_dimensions[column_letter].width = max_length + 2

print("\nRoute Summary:")
print(results_df.groupby('route_type')['passes'].sum())
print("\nTotal Routes with POI Matches:", results_df['passes'].sum())
print("Total Routes without POI Matches:", len(results_df) - results_df['passes'].sum())

# Print detailed breakdown by route type
print("\nDetailed Breakdown by Route Type:")
route_types = results_df.groupby('route_type')
for route_type in route_types.groups:
    routes_of_type = route_types.get_group(route_type)
    matches = routes_of_type['passes'].sum()
    total = len(routes_of_type)
    no_matches = total - matches
    print(f"{route_type}:")
    print(f"  - With POI matches: {matches}")
    print(f"  - Without POI matches: {no_matches}")
    print(f"  - Total routes: {total}")

# Visualization with enhanced Folium mapping
m = folium.Map(location=(site_lat, site_lon), zoom_start=12)

# Color mapping for different route types
color_map = {
    'origin_to_site': 'blue',
    'site_to_destination': 'green',
    'site_to_site_1': 'red',
    'site_to_site_2': 'purple'
}

# Add site zone marker
folium.Marker(
    location=(site_lat, site_lon),
    popup=f"Site Zone {site_zone}",
    icon=folium.Icon(color='black', icon='home')
).add_to(m)

# Add routes to the map
for index, row in results_df.iterrows():
    if index >= len(routes):  # Skip if no route was generated (error cases)
        continue
        
    route_geometry = routes[index]
    
    if route_geometry:  # Only process non-None routes
        route_coords = decode(route_geometry)
        
        route_color = color_map.get(row['route_type'], 'gray')
        folium.PolyLine(
            route_coords, 
            color=route_color, 
            weight=5, 
            opacity=0.7,
            popup=f"Route Type: {row['route_type']}<br>Origin: {row['origin_id']}<br>Destination: {row['dest_id']}"
        ).add_to(m)

# Add POIs with tooltips
for poi in pois:
    folium.CircleMarker(
        location=poi['coordinates'],
        radius=5,
        popup=f"POI ID: {poi['id']}<br>Name: {poi['name']}<br>Coordinates: {poi['coordinates']}<br>Threshold: {poi['threshold']} km",
        color='orange',
        fill=True,
        fillColor='orange',
        fillOpacity=0.7
    ).add_to(m)

    # Add proximity circle for POIs with individual thresholds
    folium.Circle(
        location=poi['coordinates'],
        radius=poi['threshold'] * 1000,  # Convert km to meters
        color='orange',
        fill=True,
        fillOpacity=0.2,
        popup=f"POI Threshold: {poi['threshold']} km"
    ).add_to(m)

# Add markers for origins and destinations that passed through POIs
for index, row in results_df[results_df['passes']].iterrows():
    # Get the corresponding zone coordinates
    origin_zone = zones_df[zones_df['GTA06'] == row['origin_id']]
    dest_zone = zones_df[zones_df['GTA06'] == row['dest_id']]
    
    # Add markers for zones with POI matches
    if not origin_zone.empty:
        folium.Marker(
            location=(origin_zone['Latitude'].values[0], origin_zone['Longitude'].values[0]),
            popup=f"Origin Zone {row['origin_id']} with POI Match",
            icon=folium.Icon(color='red', icon='info-sign')
        ).add_to(m)
    
    if not dest_zone.empty:
        folium.Marker(
            location=(dest_zone['Latitude'].values[0], dest_zone['Longitude'].values[0]),
            popup=f"Destination Zone {row['dest_id']} with POI Match",
            icon=folium.Icon(color='purple', icon='info-sign')
        ).add_to(m)

# Add a layer control to toggle different route types
folium.LayerControl().add_to(m)

# Save the map with additional features
m.save("Route_map.html")
print("Map saved.")
